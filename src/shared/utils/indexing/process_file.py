from io import BytesIO
from docx import Document
from pypdf import PdfReader
from typing import List
from openpyxl import load_workbook


def load_pdf_from_file(pdf_file: BytesIO) -> List[str]:
    """
    Đọc nội dung từ file PDF.

    Args:
        pdf_file (BytesIO): File PDF được truyền vào dưới dạng BytesIO.

    Returns:
        List[str]: Danh sách chứa nội dung văn bản từ các trang PDF.
    """
    reader = PdfReader(pdf_file)
    text = "\n".join(
        [page.extract_text() for page in reader.pages if page.extract_text()]
    )
    return [text]


def load_docx_from_file(docx_file: BytesIO) -> List[str]:
    """
    Đọc nội dung từ file DOCX.

    Args:
        docx_file (BytesIO): File DOCX được truyền vào dưới dạng BytesIO.

    Returns:
        List[str]: Danh sách chứa nội dung văn bản từ các đoạn (paragraphs) của file DOCX.
    """
    doc = Document(docx_file)
    text = "\n".join([para.text for para in doc.paragraphs])
    return [text]


def load_txt_from_file(txt_file: BytesIO) -> List[str]:
    """
    Đọc nội dung từ file TXT.

    Args:
        txt_file (BytesIO): File TXT được truyền vào dưới dạng BytesIO.

    Returns:
        List[str]: Danh sách chứa nội dung văn bản từ file TXT.
    """
    return [txt_file.read().decode("utf-8")]


def load_excel_from_file(excel_file: BytesIO) -> List[str]:
    """
    Đọc nội dung từ file Excel (chỉ lấy dữ liệu từ sheet đầu tiên).

    Args:
        excel_file (BytesIO): File Excel được truyền vào dưới dạng BytesIO.

    Returns:
        List[str]: Danh sách chứa dữ liệu dưới dạng chuỗi CSV của sheet đầu tiên.
    """
    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    docs = "\n".join([str(row) for row in data])
    return [docs]
